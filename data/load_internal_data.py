"""
load_internal_data.py
=====================
Loads two internal Excel files into PostgreSQL:

  1. Yarn cost records  → lkp_yarn_taxonomy + yarn_costs
     Source: EMİNE BİRİM MALİYET KDV SİZ-SON .xls  (sheet: İPLİK )

  2. Completed orders   → orders + order_invoices
     Source: siparis_biten_CLEANED_expert_corrected.xlsx
             sheets: siparis_biten_clean + fatura_detay

Run:
    python data/load_internal_data.py

Environment:
    DATABASE_URL — PostgreSQL connection string (required)
"""

import os
import sys
import re
import logging
from datetime import date, datetime

import xlrd
import openpyxl
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv

load_dotenv()

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# File paths
# ---------------------------------------------------------------------------
DOWNLOADS = os.path.join(os.path.expanduser("~"), "Downloads")

YARN_FILE = os.path.join(DOWNLOADS, "EMİNE BİRİM MALİYET KDV SİZ-SON .xls")
ORDERS_FILE = os.path.join(DOWNLOADS, "siparis_biten_CLEANED_expert_corrected.xlsx")

YARN_SHEET = "İPLİK "          # trailing space is intentional
ORDERS_SHEET = "siparis_biten_clean"
INVOICES_SHEET = "fatura_detay"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def to_float(val):
    """Convert a cell value to float, handling Turkish decimal format (1.234,56)."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val) if not (isinstance(val, float) and (val != val)) else None  # NaN guard
    s = str(val).strip()
    if not s:
        return None
    # Turkish format: dots as thousands sep, comma as decimal
    # e.g. "1.234,56" → "1234.56"
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def to_date_xlrd(val, datemode):
    """Convert an xlrd cell to a Python date (handles Excel date serials and strings)."""
    if val is None:
        return None
    if isinstance(val, float):
        try:
            return xlrd.xldate_as_datetime(val, datemode).date()
        except Exception:
            return None
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def to_date_openpyxl(val):
    """Convert an openpyxl cell value to a Python date."""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.date() if hasattr(val, "date") else val
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def to_text(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


# ---------------------------------------------------------------------------
# Yarn cost loader
# ---------------------------------------------------------------------------

def load_yarn_data(conn):
    """
    Reads the İPLİK sheet and loads lkp_yarn_taxonomy + yarn_costs.
    Returns (taxonomy_inserted, yarn_inserted, yarn_skipped).
    """
    log.info("Opening yarn cost file: %s", YARN_FILE)
    wb = xlrd.open_workbook(YARN_FILE)

    # Find the sheet (name has trailing space)
    sheet = None
    for name in wb.sheet_names():
        if name.strip() == "İPLİK":
            sheet = wb.sheet_by_name(name)
            break
    if sheet is None:
        raise RuntimeError(f"Sheet 'İPLİK' not found. Available: {wb.sheet_names()}")

    log.info("Sheet '%s': %d rows x %d cols", sheet.name, sheet.nrows, sheet.ncols)

    # --- Discover header row ---
    # Look for a row whose cells contain recognizable Turkish column names
    header_row_idx = None
    col_map = {}  # logical_name → col_index

    HEADER_KEYWORDS = {
        "FABR": "factory_entry_date",
        "KUMAŞ CİNSİ": "yarn_type",
        "KUTU": "boxes",
        "KG": "qty_kg",
        "ALPER MASRAFLARI TL": "cost_tl",
        "KUR": "exchange_rate",
        "ALPER MASRAFI USD KARŞILIĞI": "cost_usd_ratio",
        "ALPER MASRAFLARI USD": "cost_usd",
        "TOPLAM MASRAF USD": "total_cost_usd",
        "FATURA USD TUTARI": "invoice_usd",
        "HAM MALİYET": "unit_cost_usd",
        "FATURA NO": "invoice_no",
        "FATURA TARİHİ": "invoice_date",
        "FİYAT": "price",
        "KİMDEN": "supplier",
    }

    for ridx in range(min(10, sheet.nrows)):
        row_vals = [str(sheet.cell_value(ridx, c)).strip().upper() for c in range(sheet.ncols)]
        matched = 0
        for kw in HEADER_KEYWORDS:
            for cidx, cell_str in enumerate(row_vals):
                if kw.upper() in cell_str:
                    matched += 1
                    break
        if matched >= 5:
            header_row_idx = ridx
            # Build col_map
            for cidx, cell_str in enumerate(row_vals):
                for kw, logical in HEADER_KEYWORDS.items():
                    if kw.upper() in cell_str and logical not in col_map:
                        col_map[logical] = cidx
            break

    if header_row_idx is None:
        raise RuntimeError("Could not find header row in İPLİK sheet")
    log.info("Header at row %d; columns: %s", header_row_idx, col_map)

    datemode = wb.datemode

    # --- Collect rows ---
    yarn_rows = []
    for ridx in range(header_row_idx + 1, sheet.nrows):
        def cell(key):
            cidx = col_map.get(key)
            if cidx is None:
                return None
            c = sheet.cell(ridx, cidx)
            return c.value if c.ctype != xlrd.XL_CELL_EMPTY else None

        yarn_type = to_text(cell("yarn_type"))
        if not yarn_type:
            continue  # skip empty / sub-total rows

        # parse date-type cells
        def cell_date(key):
            cidx = col_map.get(key)
            if cidx is None:
                return None
            c = sheet.cell(ridx, cidx)
            if c.ctype == xlrd.XL_CELL_DATE:
                try:
                    return xlrd.xldate_as_datetime(c.value, datemode).date()
                except Exception:
                    return None
            if c.ctype == xlrd.XL_CELL_EMPTY:
                return None
            return to_date_xlrd(c.value, datemode)

        yarn_rows.append({
            "factory_entry_date": cell_date("factory_entry_date"),
            "yarn_type":          yarn_type,
            "boxes":              to_float(cell("boxes")),
            "qty_kg":             to_float(cell("qty_kg")),
            "cost_tl":            to_float(cell("cost_tl")),
            "exchange_rate":      to_float(cell("exchange_rate")),
            "cost_usd_ratio":     to_float(cell("cost_usd_ratio")),
            "cost_usd":           to_float(cell("cost_usd")),
            "total_cost_usd":     to_float(cell("total_cost_usd")),
            "invoice_usd":        to_float(cell("invoice_usd")),
            "unit_cost_usd":      to_float(cell("unit_cost_usd")),
            "invoice_no":         to_text(cell("invoice_no")),
            "invoice_date":       cell_date("invoice_date"),
            "price":              to_float(cell("price")),
            "supplier":           to_text(cell("supplier")),
            "source_file":        os.path.basename(YARN_FILE),
        })

    log.info("Parsed %d yarn cost rows", len(yarn_rows))

    # --- Derive taxonomy ---
    distinct_types = sorted({r["yarn_type"] for r in yarn_rows if r["yarn_type"]})
    log.info("Distinct yarn types: %d", len(distinct_types))

    cur = conn.cursor()

    # Load lkp_yarn_taxonomy
    taxonomy_inserted = 0
    for yt in distinct_types:
        cur.execute(
            """
            INSERT INTO lkp_yarn_taxonomy (yarn_type)
            VALUES (%s)
            ON CONFLICT (yarn_type) DO NOTHING
            """,
            (yt,),
        )
        taxonomy_inserted += cur.rowcount

    conn.commit()
    log.info("lkp_yarn_taxonomy: %d inserted", taxonomy_inserted)

    # Load yarn_costs (truncate + reload for idempotency)
    cur.execute("TRUNCATE yarn_costs RESTART IDENTITY CASCADE")
    psycopg2.extras.execute_batch(
        cur,
        """
        INSERT INTO yarn_costs (
            factory_entry_date, yarn_type, boxes, qty_kg,
            cost_tl, exchange_rate, cost_usd_ratio, cost_usd,
            total_cost_usd, invoice_usd, unit_cost_usd,
            invoice_no, invoice_date, price, supplier, source_file
        ) VALUES (
            %(factory_entry_date)s, %(yarn_type)s, %(boxes)s, %(qty_kg)s,
            %(cost_tl)s, %(exchange_rate)s, %(cost_usd_ratio)s, %(cost_usd)s,
            %(total_cost_usd)s, %(invoice_usd)s, %(unit_cost_usd)s,
            %(invoice_no)s, %(invoice_date)s, %(price)s, %(supplier)s, %(source_file)s
        )
        """,
        yarn_rows,
        page_size=200,
    )
    conn.commit()
    yarn_inserted = len(yarn_rows)
    log.info("yarn_costs: %d rows inserted", yarn_inserted)

    cur.close()
    return taxonomy_inserted, yarn_inserted


# ---------------------------------------------------------------------------
# Orders loader
# ---------------------------------------------------------------------------

def _wb_header_map(sheet, expected_keywords):
    """
    Find header row in an openpyxl worksheet.
    Returns (header_row_idx, col_map) where col_map maps
    logical_name -> 0-based column index.
    """
    for ridx, row in enumerate(sheet.iter_rows(max_row=10, values_only=True)):
        row_upper = [str(v).strip().upper() if v is not None else "" for v in row]
        matched = sum(1 for kw in expected_keywords if any(kw in cell for cell in row_upper))
        if matched >= 3:
            col_map = {}
            for kw, logical in expected_keywords.items():
                for cidx, cell_str in enumerate(row_upper):
                    if kw in cell_str and logical not in col_map:
                        col_map[logical] = cidx
            return ridx, col_map
    return None, {}


def load_orders_data(conn):
    """
    Reads siparis_biten_clean + fatura_detay sheets and loads orders + order_invoices.
    Returns (orders_inserted, invoices_inserted).
    """
    log.info("Opening orders file: %s", ORDERS_FILE)
    wb = openpyxl.load_workbook(ORDERS_FILE, read_only=True, data_only=True)

    # ---- ORDERS sheet ----
    if ORDERS_SHEET not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{ORDERS_SHEET}' not found. Available: {wb.sheetnames}")
    ws_orders = wb[ORDERS_SHEET]

    # Read all rows into memory (read_only mode requires single pass)
    all_rows = list(ws_orders.iter_rows(values_only=True))
    log.info("Orders sheet: %d raw rows", len(all_rows))

    # Find header row
    ORDERS_KEYWORDS = {
        "ORDER_ID": "order_id",
        "ORDER_DATE": "order_date",
        "SUPPLIER_RAW": "supplier_raw",
        "SUPPLIER_CLEAN": "supplier_clean",
        "ITEM_RAW": "item_raw",
        "PRODUCT_CLEAN": "product_clean",
        "PRODUCT_GROUP_EXPERT": "product_group_expert",
        "PRODUCT_GROUP": "product_group",
        "ORDERED_QTY": "ordered_qty",
        "QTY_NUMERIC": "qty_numeric",
        "UNIT_RAW": "unit_raw",
        "UNIT_STANDARD": "unit_standard",
        "INVOICE_DATE": "invoice_date",
        "INVOICE_NO": "invoice_no",
        "RECEIVED_QTY": "received_qty",
        "REMAINING_QTY": "remaining_qty",
        "UNIT_PRICE_RAW": "unit_price_raw",
        "PRICE_NUMERIC": "price_numeric",
        "CURRENCY_RAW": "currency_raw",
        "CURRENCY_CLEAN": "currency_clean",
        "TRANSACTION_TYPE": "transaction_type",
        "RECORD_STATUS": "record_status",
        "NOTES_CLEANING": "notes_cleaning",
    }

    header_row_idx = None
    col_map = {}
    for ridx, row in enumerate(all_rows[:10]):
        row_upper = [str(v).strip().upper() if v is not None else "" for v in row]
        matched = sum(1 for kw in ORDERS_KEYWORDS if any(kw in cell for cell in row_upper))
        if matched >= 5:
            header_row_idx = ridx
            for kw, logical in ORDERS_KEYWORDS.items():
                for cidx, cell_str in enumerate(row_upper):
                    if kw in cell_str and logical not in col_map:
                        col_map[logical] = cidx
            break

    if header_row_idx is None:
        # Fallback: use first row as header
        header_row_idx = 0
        row_upper = [str(v).strip().upper() if v is not None else "" for v in all_rows[0]]
        for kw, logical in ORDERS_KEYWORDS.items():
            for cidx, cell_str in enumerate(row_upper):
                if kw in cell_str and logical not in col_map:
                    col_map[logical] = cidx

    log.info("Orders header at row %d; mapped %d columns", header_row_idx, len(col_map))

    def get(row, key):
        cidx = col_map.get(key)
        if cidx is None or cidx >= len(row):
            return None
        return row[cidx]

    order_rows = []
    seen_ids = set()
    for row in all_rows[header_row_idx + 1:]:
        order_id = to_text(get(row, "order_id"))
        if not order_id:
            continue
        if order_id in seen_ids:
            log.warning("Duplicate order_id '%s' — skipping", order_id)
            continue
        seen_ids.add(order_id)
        order_rows.append({
            "order_id":             order_id,
            "order_date":           to_date_openpyxl(get(row, "order_date")),
            "supplier_raw":         to_text(get(row, "supplier_raw")),
            "supplier_clean":       to_text(get(row, "supplier_clean")),
            "item_raw":             to_text(get(row, "item_raw")),
            "product_clean":        to_text(get(row, "product_clean")),
            "product_group":        to_text(get(row, "product_group")),
            "product_group_expert": to_text(get(row, "product_group_expert")),
            "ordered_qty":          to_float(get(row, "ordered_qty")),
            "qty_numeric":          to_float(get(row, "qty_numeric")),
            "unit_raw":             to_text(get(row, "unit_raw")),
            "unit_standard":        to_text(get(row, "unit_standard")),
            "invoice_date":         to_date_openpyxl(get(row, "invoice_date")),
            "invoice_no":           to_text(get(row, "invoice_no")),
            "received_qty":         to_float(get(row, "received_qty")),
            "remaining_qty":        to_float(get(row, "remaining_qty")),
            "unit_price_raw":       to_text(get(row, "unit_price_raw")),
            "price_numeric":        to_float(get(row, "price_numeric")),
            "currency_raw":         to_text(get(row, "currency_raw")),
            "currency_clean":       to_text(get(row, "currency_clean")),
            "transaction_type":     to_text(get(row, "transaction_type")),
            "record_status":        to_text(get(row, "record_status")),
            "notes_cleaning":       to_text(get(row, "notes_cleaning")),
        })

    log.info("Parsed %d order rows", len(order_rows))

    # ---- INVOICES sheet ----
    if INVOICES_SHEET not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{INVOICES_SHEET}' not found. Available: {wb.sheetnames}")
    ws_inv = wb[INVOICES_SHEET]
    all_inv_rows = list(ws_inv.iter_rows(values_only=True))
    log.info("Invoices sheet: %d raw rows", len(all_inv_rows))

    INV_KEYWORDS = {
        "PARENT_ORDER_ID": "parent_order_id",
        "INVOICE_DATE": "invoice_date",
        "INVOICE_NO": "invoice_no",
        "UNIT_PRICE": "unit_price",
        "CURRENCY": "currency",
        "RECEIVED_QTY": "received_qty",
        "UNIT": "unit",
        "TRANSACTION_TYPE": "transaction_type",
        "NOTES": "notes",
        "PRODUCT_GROUP_EXPERT": "product_group_expert",
        "PRODUCT_GROUP_REVIEW": "product_group_review",
    }

    inv_header_idx = None
    inv_col_map = {}
    for ridx, row in enumerate(all_inv_rows[:10]):
        row_upper = [str(v).strip().upper() if v is not None else "" for v in row]
        matched = sum(1 for kw in INV_KEYWORDS if any(kw in cell for cell in row_upper))
        if matched >= 3:
            inv_header_idx = ridx
            for kw, logical in INV_KEYWORDS.items():
                for cidx, cell_str in enumerate(row_upper):
                    if kw in cell_str and logical not in inv_col_map:
                        inv_col_map[logical] = cidx
            break

    if inv_header_idx is None:
        inv_header_idx = 0
        row_upper = [str(v).strip().upper() if v is not None else "" for v in all_inv_rows[0]]
        for kw, logical in INV_KEYWORDS.items():
            for cidx, cell_str in enumerate(row_upper):
                if kw in cell_str and logical not in inv_col_map:
                    inv_col_map[logical] = cidx

    log.info("Invoices header at row %d; mapped %d columns", inv_header_idx, len(inv_col_map))

    def getinv(row, key):
        cidx = inv_col_map.get(key)
        if cidx is None or cidx >= len(row):
            return None
        return row[cidx]

    valid_order_ids = {r["order_id"] for r in order_rows}
    inv_rows = []
    inv_skipped = 0
    for row in all_inv_rows[inv_header_idx + 1:]:
        parent_id = to_text(getinv(row, "parent_order_id"))
        if not parent_id:
            continue
        if parent_id not in valid_order_ids:
            inv_skipped += 1
            continue
        inv_rows.append({
            "parent_order_id":      parent_id,
            "invoice_date":         to_date_openpyxl(getinv(row, "invoice_date")),
            "invoice_no":           to_text(getinv(row, "invoice_no")),
            "unit_price":           to_float(getinv(row, "unit_price")),
            "currency":             to_text(getinv(row, "currency")),
            "received_qty":         to_float(getinv(row, "received_qty")),
            "unit":                 to_text(getinv(row, "unit")),
            "transaction_type":     to_text(getinv(row, "transaction_type")),
            "notes":                to_text(getinv(row, "notes")),
            "product_group_expert": to_text(getinv(row, "product_group_expert")),
            "product_group_review": to_text(getinv(row, "product_group_review")),
        })

    log.info("Parsed %d invoice rows (%d skipped — orphan parent_order_id)", len(inv_rows), inv_skipped)

    # ---- Write to DB ----
    cur = conn.cursor()

    # Truncate + reload orders (idempotent)
    cur.execute("TRUNCATE order_invoices, orders RESTART IDENTITY CASCADE")
    conn.commit()

    psycopg2.extras.execute_batch(
        cur,
        """
        INSERT INTO orders (
            order_id, order_date, supplier_raw, supplier_clean,
            item_raw, product_clean, product_group, product_group_expert,
            ordered_qty, qty_numeric, unit_raw, unit_standard,
            invoice_date, invoice_no, received_qty, remaining_qty,
            unit_price_raw, price_numeric, currency_raw, currency_clean,
            transaction_type, record_status, notes_cleaning
        ) VALUES (
            %(order_id)s, %(order_date)s, %(supplier_raw)s, %(supplier_clean)s,
            %(item_raw)s, %(product_clean)s, %(product_group)s, %(product_group_expert)s,
            %(ordered_qty)s, %(qty_numeric)s, %(unit_raw)s, %(unit_standard)s,
            %(invoice_date)s, %(invoice_no)s, %(received_qty)s, %(remaining_qty)s,
            %(unit_price_raw)s, %(price_numeric)s, %(currency_raw)s, %(currency_clean)s,
            %(transaction_type)s, %(record_status)s, %(notes_cleaning)s
        )
        """,
        order_rows,
        page_size=200,
    )
    conn.commit()
    orders_inserted = len(order_rows)
    log.info("orders: %d rows inserted", orders_inserted)

    if inv_rows:
        psycopg2.extras.execute_batch(
            cur,
            """
            INSERT INTO order_invoices (
                parent_order_id, invoice_date, invoice_no,
                unit_price, currency, received_qty, unit,
                transaction_type, notes, product_group_expert, product_group_review
            ) VALUES (
                %(parent_order_id)s, %(invoice_date)s, %(invoice_no)s,
                %(unit_price)s, %(currency)s, %(received_qty)s, %(unit)s,
                %(transaction_type)s, %(notes)s, %(product_group_expert)s, %(product_group_review)s
            )
            """,
            inv_rows,
            page_size=500,
        )
        conn.commit()
    invoices_inserted = len(inv_rows)
    log.info("order_invoices: %d rows inserted", invoices_inserted)

    cur.close()
    return orders_inserted, invoices_inserted


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    db_url = os.getenv("DATABASE_URL")
    if not db_url:
        sys.exit("ERROR: DATABASE_URL not set")

    # Verify source files exist
    for path, label in [(YARN_FILE, "Yarn cost"), (ORDERS_FILE, "Orders")]:
        if not os.path.exists(path):
            sys.exit(f"ERROR: {label} file not found: {path}")

    conn = psycopg2.connect(db_url)
    try:
        taxonomy_inserted, yarn_inserted = load_yarn_data(conn)
        orders_inserted, invoices_inserted = load_orders_data(conn)
    finally:
        conn.close()

    print()
    print("=" * 50)
    print("Load complete — summary")
    print("=" * 50)
    print(f"  lkp_yarn_taxonomy : {taxonomy_inserted:>6} new types inserted")
    print(f"  yarn_costs        : {yarn_inserted:>6} rows inserted")
    print(f"  orders            : {orders_inserted:>6} rows inserted")
    print(f"  order_invoices    : {invoices_inserted:>6} rows inserted")
    print("=" * 50)


if __name__ == "__main__":
    main()
